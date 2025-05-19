import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# === Config ===
base_dir = Path(__file__).resolve().parent
prediction_dir = base_dir
match_finished_path = base_dir.parent.parent / "data" / "current" / "matches_finished.csv"

# === Load finished results
print(f"📅 Loading matches_finished.csv...")
mf = pd.read_csv(match_finished_path)
mf = mf[mf["winner_code"].isin(["A", "B"])]


def light_normalize(name):
    name = str(name)
    name = re.sub(r"\[\d+\]\d+\.\d+", "", name)
    name = re.sub(r"\[\d+\]", "", name)
    name = re.sub(r"\d+\.\d+$", "", name)
    name = re.sub(r"\([^)]*\)", "", name)
    return name.strip().lower()


mf["player_A_ln"] = mf["player_A"].apply(light_normalize)
mf["player_B_ln"] = mf["player_B"].apply(light_normalize)
mf["winner_clean"] = mf.apply(
    lambda r: r["player_A_ln"] if r["winner_code"] == "A" else r["player_B_ln"], axis=1
)

# === Resolution logic
def apply_resolution(df):
    df["player_A_ln"] = df["player_A"].apply(light_normalize)
    df["player_B_ln"] = df["player_B"].apply(light_normalize)

    def resolve(row):
        a, b = row["player_A_ln"], row["player_B_ln"]
        matches = mf[((mf["player_A_ln"] == a) & (mf["player_B_ln"] == b)) |
                     ((mf["player_A_ln"] == b) & (mf["player_B_ln"] == a))]
        if not matches.empty:
            winner = matches.iloc[0]["winner_clean"]
            return row["player_A"] if winner == a else row["player_B"]
        return None

    df["correct_pick"] = df.apply(resolve, axis=1)
    df["status"] = df.apply(
        lambda r: "🟢 CAPPED" if r["capbot_pick"] == r["correct_pick"]
        else ("🔴 FLOP" if pd.notnull(r["correct_pick"]) else "⏳ Awaiting Result"), axis=1
    )

    df.drop(columns=["player_A_ln", "player_B_ln"], inplace=True, errors="ignore")
    return df

# === Process All Prediction Files
files = sorted(prediction_dir.glob("CapBot_v1d_Predictions_*.xlsx"))
print(f"🔍 Found {len(files)} prediction files...")

for file in files:
    print(f"\n📄 Processing: {file.name}")
    try:
        wb = load_workbook(file)
        available_sheets = wb.sheetnames
    except Exception as e:
        print(f"❌ Failed to open {file.name}: {e}")
        continue

    sheets_to_update = ["Full Predictions", "Filtered Picks", "Filtered FLOPS"]
    existing_targets = [s for s in sheets_to_update if s in available_sheets]
    updated_sheets = {}

    for sheet in existing_targets:
        try:
            df = pd.read_excel(file, sheet_name=sheet)

            # ✅ Separate tag legend rows BEFORE any processing
            legend_rows = pd.DataFrame()
            if "match_id" in df.columns:
                legend_rows = df[df["match_id"].isnull()].copy()
                df = df[df["match_id"].notnull()].copy()
            else:
                df = df.copy()

            print(f"  🔁 Resolving: {sheet}")
            resolved_df = apply_resolution(df)

            # ✅ Clean and re-attach tag legend
            if not legend_rows.empty:
                if "Tag" in legend_rows.columns:
                    tag_mask = legend_rows["Tag"].notnull()
                    for col in legend_rows.columns:
                        if col not in ["Tag", "Description"]:
                            legend_rows.loc[tag_mask, col] = ""

                # ✅ REAL spacer row
                spacer = pd.DataFrame([{col: None for col in resolved_df.columns}])
                resolved_df = pd.concat([resolved_df, spacer, legend_rows], ignore_index=True)

            updated_sheets[sheet] = resolved_df

        except Exception as e:
            print(f"  ⚠️ Failed to update '{sheet}': {e}")

    # === Regenerate Summary
    summary_df = None
    if "Full Predictions" in updated_sheets:
        try:
            summary_rows = [
                {
                    "Type": "Full", "Date": file.stem[-8:], "Total Matches": len(updated_sheets["Full Predictions"]),
                    "Correct": updated_sheets["Full Predictions"]["capbot_pick"].eq(updated_sheets["Full Predictions"]["correct_pick"]).sum(),
                    "Accuracy": round(updated_sheets["Full Predictions"]["capbot_pick"].eq(updated_sheets["Full Predictions"]["correct_pick"]).mean(), 4),
                    "Average EV": round(updated_sheets["Full Predictions"]["ev"].mean(), 4),
                    "Average Stake": round(updated_sheets["Full Predictions"]["stake"].mean(), 2),
                    "Average Proba": round(updated_sheets["Full Predictions"]["confidence"].mean(), 4)
                }
            ]
            if "Filtered Picks" in updated_sheets:
                filtered_df = updated_sheets["Filtered Picks"]
                filtered_real = filtered_df[filtered_df["match_id"].notnull()]
                summary_rows.append({
                    "Type": "Filtered", "Date": file.stem[-8:], "Total Matches": len(filtered_real),
                    "Correct": filtered_real["capbot_pick"].eq(filtered_real["correct_pick"]).sum(),
                    "Accuracy": round(filtered_real["capbot_pick"].eq(filtered_real["correct_pick"]).mean(), 4),
                    "Average EV": round(filtered_real["ev"].mean(), 4),
                    "Average Stake": round(filtered_real["stake"].mean(), 2),
                    "Average Proba": round(filtered_real["confidence"].mean(), 4)
                })
            summary_df = pd.DataFrame(summary_rows)
        except Exception as e:
            print(f"  ⚠️ Could not regenerate Summary: {e}")

    # === Save all updates to file
    try:
        with pd.ExcelWriter(file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            for sheet, output_df in updated_sheets.items():
                if sheet in writer.book.sheetnames:
                    std = writer.book[sheet]
                    writer.book.remove(std)
                    writer.book.create_sheet(sheet)
                output_df.to_excel(writer, sheet_name=sheet, index=False)
            if summary_df is not None:
                if "Summary" in writer.book.sheetnames:
                    std = writer.book["Summary"]
                    writer.book.remove(std)
                    writer.book.create_sheet("Summary")
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # === Final cleanup pass: remove lingering \u23f3 rows below legend
        post_wb = load_workbook(file)
        for sheet_name in ["Full Predictions", "Filtered Picks", "Filtered FLOPS"]:
            if sheet_name in post_wb.sheetnames:
                ws = post_wb[sheet_name]
                found_tag = False
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    values = [cell.value for cell in row]
                    if not found_tag and values[0] == "Tag":
                        found_tag = True
                        continue
                    if found_tag:
                        if any(isinstance(val, str) and "⏳ Awaiting Result" in val for val in values):
                            ws.delete_rows(row[0].row, 1)
        post_wb.save(file)

        print(f"✅ Finished updating: {file.name}")
    except Exception as e:
        print(f"❌ Failed to save updates to {file.name}: {e}")
